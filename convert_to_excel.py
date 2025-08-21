#!/usr/bin/env python3
"""
Script untuk mengkonversi file CSV data relawan ke Excel dengan formatting yang lebih baik
"""

import pandas as pd
import os
from datetime import datetime
import glob

def convert_csv_to_excel():
    """Konversi file CSV terbaru ke Excel dengan formatting"""
    
    # Cari file CSV terbaru
    csv_files = glob.glob("Data_Relawan_Lengkap_*.csv")
    family_csv_files = glob.glob("Detail_Keluarga_Relawan_*.csv")
    
    if not csv_files:
        print("Tidak ada file CSV data relawan yang ditemukan!")
        return
    
    # Ambil file CSV terbaru
    main_csv = max(csv_files, key=os.path.getctime)
    family_csv = max(family_csv_files, key=os.path.getctime) if family_csv_files else None
    
    print(f"Membaca file: {main_csv}")
    
    # Baca CSV dengan encoding UTF-8
    df_main = pd.read_csv(main_csv, encoding='utf-8')
    
    # Generate nama file Excel
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    excel_filename = f"Data_Relawan_Lengkap_{timestamp}.xlsx"
    
    # Buat Excel writer
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Sheet 1: Data Relawan Utama
        df_main.to_excel(writer, sheet_name='Data Relawan', index=False)
        
        # Sheet 2: Detail Keluarga (jika ada)
        if family_csv and os.path.exists(family_csv):
            print(f"📖 Membaca file keluarga: {family_csv}")
            df_family = pd.read_csv(family_csv, encoding='utf-8')
            df_family.to_excel(writer, sheet_name='Detail Keluarga', index=False)
        
        # Sheet 3: Ringkasan Statistik
        create_summary_sheet(df_main, writer)
        
        # Format sheets
        format_excel_sheets(writer, df_main)
    
    print(f"✅ File Excel berhasil dibuat: {excel_filename}")
    print(f"📊 Total relawan: {len(df_main)}")
    
    # Hitung statistik
    active_volunteers = len(df_main[df_main['Status'] == 'Aktif'])
    with_profile = len(df_main[df_main['NIK'].notna()])
    with_phone = len(df_main[df_main['Telepon'].notna()])
    
    print(f"🟢 Relawan aktif: {active_volunteers}")
    print(f"👤 Memiliki profil lengkap: {with_profile}")
    print(f"📱 Memiliki nomor telepon: {with_phone}")

def create_summary_sheet(df, writer):
    """Buat sheet ringkasan statistik"""
    
    summary_data = {
        'Kategori': [
            'Total Relawan',
            'Relawan Aktif', 
            'Relawan Tidak Aktif',
            'Memiliki Profil Lengkap (NIK)',
            'Memiliki Nomor Telepon',
            'Memiliki Data Ekonomi',
            'Memiliki Data Sosial',
            'Memiliki Anggota Keluarga'
        ],
        'Jumlah': [
            len(df),
            len(df[df['Status'] == 'Aktif']),
            len(df[df['Status'] == 'Tidak Aktif']),
            len(df[df['NIK'].notna()]),
            len(df[df['Telepon'].notna()]),
            len(df[df['Penghasilan Bulanan'].notna()]),
            len(df[df['Organisasi'].notna()]),
            len(df[df['Jumlah Anggota Keluarga'] > 0])
        ],
        'Persentase': []
    }
    
    total = len(df)
    for jumlah in summary_data['Jumlah']:
        persen = (jumlah / total * 100) if total > 0 else 0
        summary_data['Persentase'].append(f"{persen:.1f}%")
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Ringkasan Statistik', index=False)

def format_excel_sheets(writer, df_main):
    """Format Excel sheets untuk tampilan yang lebih baik"""
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Format sheet Data Relawan
    ws_main = writer.sheets['Data Relawan']
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for col in range(1, len(df_main.columns) + 1):
        cell = ws_main.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in ws_main.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max width 50
        ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws_main.iter_rows(min_row=1, max_row=len(df_main) + 1, 
                                 min_col=1, max_col=len(df_main.columns)):
        for cell in row:
            cell.border = thin_border
    
    # Zebra striping for data rows
    light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    for row_num in range(2, len(df_main) + 2):
        if row_num % 2 == 0:
            for col in range(1, len(df_main.columns) + 1):
                ws_main.cell(row=row_num, column=col).fill = light_fill
    
    print("🎨 Format Excel berhasil diterapkan")

if __name__ == "__main__":
    try:
        convert_csv_to_excel()
    except ImportError:
        print("❌ Error: pandas dan openpyxl diperlukan!")
        print("💡 Install dengan: pip install pandas openpyxl")
    except Exception as e:
        print(f"❌ Error: {e}")